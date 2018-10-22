import openpyxl
import sys
from twilio.rest import Client
from datetime import datetime
import pandas as pd
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from os.path import basename
import xlsxwriter
from tkinter import *
from tkinter import messagebox
from pandastable import Table, TableModel


account_sid = 'ACffd2fce359a976ec9b7eb4544bb503fd'
auth_token = '8907db481180c1ff11a26faaba44746e'
client = Client(account_sid, auth_token)

username = ''
phoneNum = ''
email = ''
spendLimit = ''
loginSuccess = False


def resetValues():
    price.delete(0,END)
    item.delete(0,END)
    date.delete(0,END)

def register1():
    global username
    username = username_value.get()
    file = open('passwd.txt', 'r')
    taken = False
    userMatch = []
    for line in file:
        userMatch = line.split(':')
        print(userMatch)
        if (userMatch[0] == username):
            taken = True
            break
    if (taken == True):
    	messagebox.showerror("Username Taken", "The username has been taken, please try again.")
        #print("Sorry, that username has been taken. Please try again")
        
    else:
        global loginSuccess
        global phoneNum
        global email
        global spendLimit
        password = password_value.get()
        phoneNum = phonenumber_value.get()
        email = email_value.get()
        spendLimit = limit_value.get()
        file = open('passwd.txt', 'a')
        file.write(username)
        file.write(":")
        file.write(password)
        file.write(":")
        file.write(phoneNum)
        file.write(":")
        file.write(email)
        file.write(":")
        file.write(spendLimit)
        file.write("\n")
        file.close()
        phoneNum = userMatch[2]
        email = userMatch[3]
        spendLimit = userMatch[4]
        loginSuccess = True
        workbook = xlsxwriter.Workbook(username + '.xlsx')
        worksheet = workbook.add_worksheet()
        workbook.close()
        #mainMenu()

def login1():
    global loginSuccess
    global phoneNum
    global email
    global spendLimit
    global username
    print("Please Log In")
    username = loginUsername_value.get()
    password = loginPassword_value.get()
    login_info = []
    loginSuccess = False
    for line in open('passwd.txt', 'r'):
        login_info = line.split(':')
        if (username == login_info[0] and password == login_info[1]):
            messagebox.showinfo("Success", "Hello " + username + ", you are now successfully logged in.")
            
            username = login_info[0]
            phoneNum = login_info[2]
            email = login_info[3]
            spendLimit = login_info[4]
            loginSuccess = True
            #mainMenu()
    if loginSuccess == False:
        messagebox.showerror("Invalid Login", "Incorrect credentials. Please try again.")
        loginWindow()


def recordExpense1():
    item = item_value.get()
    print(item)
    price = price_value.get()
    date = date_value.get()
    wb = openpyxl.load_workbook(username + ".xlsx")
    ws = wb.worksheets[0]
    dateHeader = ws.cell(row=1, column=1)
    itemHeader = ws.cell(row=1, column=2)
    costHeader = ws.cell(row=1, column=3)
    dateHeader.value = 'Date'
    itemHeader.value = 'Item'
    costHeader.value = 'Cost'
    newRow = ws.max_row + 1
    dateRow = ws.cell(row=newRow, column=1)
    itemRow = ws.cell(row=newRow, column=2)
    costRow = ws.cell(row=newRow, column=3)
    dateRow.value = str(date)
    itemRow.value = str(item)
    costRow.value = str(price)
    wb.save(username + ".xlsx")
    messagebox.showinfo("Success", "Expenses recorded.")
    '''done = False
    while (done == False):
        response = input("Your purchase has been recorded. Record another purchase? (Y/N): ")
        if (response == 'Y'):
            done = True
            recordExpense()
        elif (response == 'N'):
            done = True
            wb.save(username + ".xlsx")
            global loginSuccess
            loginSuccess = True
            mainMenu()
        else:
            print("please enter a valid response.")
            done = False
'''
def sendEmail(send_from: str, subject: str, text: str, send_to: list, files=None):

    user = 'zee_hussain@utexas.edu'
    password = '#Buttley4Lyfe'

    send_to= send_to

    msg = MIMEMultipart()
    msg['From'] = send_from
    msg['To'] = ', '.join(send_to)
    msg['Subject'] = subject

    msg.attach(MIMEText(text))

    for f in files or []:
        with open(f, "rb") as fil:
            ext = f.split('.')[-1:]
            attachedfile = MIMEApplication(fil.read(), _subtype = ext)
            attachedfile.add_header(
                'content-disposition', 'attachment', filename=basename(f) )
        msg.attach(attachedfile)

    date = datetime.today()
    firstDate = date.today().replace(day=1)

    if (date == firstDate):
        smtp = smtplib.SMTP(host="smtp.gmail.com", port=587)
        smtp.starttls()
        smtp.login(username, password)
        smtp.sendmail(send_from, send_to, msg.as_string())
        smtp.close()

def sendText():
    wb = openpyxl.load_workbook(username + ".xlsx")
    ws = wb.worksheets[0]
    global spendLimit
    sum = 0
    for i in range(2,ws.max_row+1):
        costRow = ws.cell(row=i, column=3)
        sum = sum + float(costRow.value)
    if sum > float(spendLimit):
        message = client.messages \
            .create(
            body="You have exceeded your monthly spending limit of "+spendLimit,
            from_='+19495229115',
            to='+1'+phoneNum
        )

def viewExpenses():
    viewExpense = Toplevel()
    f=Frame(viewExpense)
    f.pack(fill=BOTH, expand=1)
    df = pd.read_excel(username + '.xlsx', index=False)
    #with pd.option_context('display.max_rows', None, 'display.max_columns', None):
     #   print(df)
      #  print("\n")
    table = pt = Table(f, dataframe=df,
                                showtoolbar=True, showstatusbar=True)
    pt.show()
    #mainMenu()
    viewExpense.mainloop()

def mainMenu():
    global loginSuccess
    if loginSuccess == True:
        print("Main Menu:")
        print("Please select an option by pressing the corresponding button")
        print("\n")
        print("To record an expense, please press: 1")
        print("To view your monthly expense report: 2")
        print("To exit: 3")
        print("\n")
        user = 'zee_hussain@utexas.edu'
        global username
        report = username + '.xlsx'
        sendEmail(send_from=user, subject="Monthly Expense Report",
                  text="Dear " + username + ",\n\nHere is your monthly expense report.\n\nBest,\n\nExpenseBot",
                  send_to=[email], files=[report])
        sendText()
        optionSelect = input()
        if optionSelect == '1':
            recordExpense()
        elif optionSelect == '2':
            viewExpenses()
        elif optionSelect == '3':
            sys.exit(0)

    else:
        checkUser = input("Welcome to Expense Tracker. Press R to register, and L to login: ")
        if (checkUser == 'R'):
            register()
            loginSuccess = True
        elif (checkUser == 'L'):
            login()
        else:
            mainMenu()

def main():
    mainMenu()
def checkIsInteger(x):
    try:
        int(x)
        return True
    except ValueError:
        return False

def checkTitleNoInt(x):
    testString = str(x)
    for i in x:
        if i.isdigit() == True:
            return False
            break
    else:
        i+=1

def recordExpenseWindow():
    recordExpense = Toplevel()
    recordExpense.geometry("500x500")
    Label(recordExpense, text = "Name of Item: ").grid(row = 0)
    Label(recordExpense, text = "Price of item: ").grid(row = 1)
    Label(recordExpense, text = "Date of purchase: ").grid(row = 2)
    #when you create the textbox, you need to define the variable to link it to in order to retrieve it
    item = Entry(recordExpense, textvariable = item_value)
    price = Entry(recordExpense, textvariable = price_value)
    date = Entry(recordExpense, textvariable = date_value)

    item.config(state = NORMAL)
    price.config(state = NORMAL)
    item.grid(row = 0, column = 1)
    price.grid(row = 1, column = 1)
    date.grid(row = 2, column = 1)
    
    dateButton = Button(recordExpense, text = "", command = date.delete(0,END))
    dateButton.grid(row = 5, column = 2)

    itemButton = Button(recordExpense, text = "", command = item.delete(0,END))
    itemButton.grid(row = 5, column = 2)

    priceButton = Button(recordExpense, text = "", command = price.delete(0,END))
    priceButton.grid(row = 5, column = 2)

    saveButton = Button(recordExpense, text = "Save", fg = "Blue", bg = "Grey", command = recordExpense1)
    saveButton.grid(row = 5, column = 2)

    saveButton = Button(recordExpense, text = "Exit", fg = "Blue", bg = "Grey", command = recordExpense.destroy)
    saveButton.grid(row = 5, column = 3)
    
    recordExpense.mainloop()

def close_window():
    window.destroy()

def registerWindow():
    register = Toplevel()
    register.geometry("500x500")

    Label(register, text = "Username: ").grid(row = 0)
    username = Entry(register, textvariable = username_value)
    username.grid(row = 0, column = 1)

    Label(register, text = "Password: ").grid(row = 1)
    password = Entry(register, textvariable = password_value)
    password.grid(row = 1, column = 1)

    Label(register, text = "Phone Number: ").grid(row = 2)
    phonenumber = Entry(register, textvariable = phonenumber_value)
    phonenumber.grid(row = 2, column = 1)

    Label(register, text = "Email: ").grid(row = 3)
    email = Entry(register, textvariable = email_value)
    email.grid(row = 3, column = 1)

    Label(register, text = "Spending limit: ").grid(row = 4)
    limit = Entry(register, textvariable = limit_value)
    limit.grid(row = 4, column = 1)

    submitButton = Button(register, text = "Save", fg = "Blue", bg = "Grey", command = register1)
    submitButton.grid(row = 5, column = 1)

    register.mainloop()

def loginWindow():
    login = Toplevel()
    login.geometry("200x200")

    Label(login, text = "Username: ").grid(row = 0)
    username = Entry(login, textvariable = loginUsername_value)
    username.grid(row = 0, column = 1)

    Label(login, text = "Password: ").grid(row = 1)
    password = Entry(login, textvariable = loginPassword_value)
    password.grid(row = 1, column = 1)
    submitButton = Button(login, text = "Save", fg = "Blue", bg = "Grey", command = login1)
    submitButton.grid(row = 2, column = 1)
    login.mainloop()


#creates a blank window
#every GUI has to have this and mainloop
mainWindow = Tk()
#sets size of window at start up
mainWindow.geometry("500x500")
#makes window unable to be resized by users
mainWindow.resizable(width = False, height = False)
#to populate the window with a label, create using Label(location, x)
#pack(fill =x) button will stretch however long window is stretched (left and right)
#replace above with y, will stretch height wise
#fill = BOTH, expand = True, will dynamically change both x and y
appTitle = Label(text = "Expense Calculator", bg = "Grey", fg = "White", height = 5, width = 20)
#when you don't care where the object is placed, but you just want it display in window
appTitle.grid(row = 0, column= 2)
#creating containers to place windows or widgets or objects in seperate areas of the entire screen  
"""topFrame = Frame(mainWindow)
topFrame.pack(side = TOP)

bottomFrame = Frame(mainWindow)
bottomFrame.pack(side = BOTTOM)"""

#defining variables
item_value = StringVar()
price_value = StringVar()
date_value = StringVar()
username_value = StringVar()
password_value = StringVar()
phonenumber_value = StringVar()
email_value = StringVar()
limit_value = StringVar()
loginUsername_value = StringVar()
loginPassword_value = StringVar()


#creation of button: Button(location, text, fg = (text color)) or bg = background color
quitButton = Button(text = "Quit", fg = "blue", bg = "Grey", command = quit, padx = 30, pady = 10)
recordButton = Button(text = "Record Expenses", fg = "red",bg= "Grey", command = recordExpenseWindow, padx = 30, pady = 10)
displayButton = Button(text = "Display Expenses", fg = "red",bg= "Grey", command = viewExpenses, padx = 30, pady = 10)
loginButton = Button(text = "Login", fg = "red", bg = "Grey", command = loginWindow, padx = 30, pady = 10)
registerButton = Button(text = "Register", fg = "red", bg = "Grey", command = registerWindow, padx = 30, pady = 10)
#by default, pack stacks objects on top of one another. 
#define side = to specify where on window you want

recordButton.grid(row = 1, column = 1)
quitButton.grid(row = 1, column = 2)
displayButton.grid(row = 1, column = 3)
loginButton.grid(row = 2, column = 1)
registerButton.grid(row = 2, column = 3)

#place window in infinite loop, keeps it displayed, until you press close
mainWindow.mainloop()

